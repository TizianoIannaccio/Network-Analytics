"""
	NETWORK ANALYTICS - 2020/2021 - Exam Project
	Tiziano Iannaccio, Mihaela Nedkova Naydenova, Gianluca Mecci
	SG
"""
import networkx as nx 
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from networkx.algorithms.approximation import clique
import networkx.algorithms.community as nx_comm
from sklearn import linear_model as lr
import math
import time

m=[] #it will contain as many of each family's name as the time members of that family appear in the graph, useful for node's color auto-assignment based on the family
color_map=[] #it will contain 5 exadecimal colors, each one is peculiar to one family. It will be repeted as many times as its family is counted in m
 

def printgood(text): 
     l= len(text)
     c = '*' * (l + 4)
     print("")
     print(c)
     print("* "+text+' *')
     print(c)



def start():
	printgood("Network Analytics: The Five Families in a Net")
	t=graph()
	G = t[0] #first element returned by graph()
	time.sleep(5)
	m=t[1] #second element returned by graph()
	for i in range(m.count("Bonanno")):
		color_map.append("#15b5e7")
	for i in range(m.count("Gambino")):
		color_map.append("#8cb800") 
	for i in range(m.count("Colombo")):
		color_map.append("#1b6f53") 
	for i in range(m.count("Lucchese")):
		color_map.append("#ff6a49") 
	for i in range(m.count("Genovese")):
		color_map.append("#915791") #useful for node's color auto-assignment based on the family
	analysis(G)



def graph():
	print("\nCreating the graph, please wait a few seconds...\n")
	print("Accessing the Node-table...")
	workbook_nodes = load_workbook(filename="spreadsheets/Nodes_list.xlsx")
	workbook_nodes.sheetnames
	nodes = workbook_nodes.active
	nodes_id=[]
	Name_dict= {}
	Nickname_dict= {}
	Boss_dict= {}
	Acting_Boss_dict= {}
	Underboss_dict= {}
	Acting_Underboss_dict= {}
	Consigliere_dict= {}
	Acting_Consigliere_dict= {}
	Capo_dict= {}
	Family_dict= {}
	twenties_dict= {}
	thirties_dict= {}
	forties_dict= {}
	fifties_dict= {}
	sixties_dict= {}
	seventies_dict= {}
	eighties_dict= {}
	print("Adding the nodes' attributes...")
	for i in range(len(nodes["A"])-1):
		Name_dict[int(nodes[i+2][0].value)]= nodes[i+2][1].value
		Nickname_dict[int(nodes[i+2][0].value)]= nodes[i+2][2].value
		Boss_dict[int(nodes[i+2][0].value)]= nodes[i+2][3].value
		Acting_Boss_dict[int(nodes[i+2][0].value)]= nodes[i+2][4].value
		Underboss_dict[int(nodes[i+2][0].value)]= nodes[i+2][5].value
		Acting_Underboss_dict[int(nodes[i+2][0].value)]= nodes[i+2][6].value
		Consigliere_dict[int(nodes[i+2][0].value)]= nodes[i+2][7].value
		Acting_Consigliere_dict[int(nodes[i+2][0].value)]= nodes[i+2][8].value
		Capo_dict[int(nodes[i+2][0].value)]= nodes[i+2][9].value
		Family_dict[int(nodes[i+2][0].value)]= nodes[i+2][10].value
		m.append(nodes[i+2][10].value)  #useful for node's color auto-assignment based on the family
		twenties_dict[int(nodes[i+2][0].value)]= nodes[i+2][11].value
		thirties_dict[int(nodes[i+2][0].value)]= nodes[i+2][12].value
		forties_dict[int(nodes[i+2][0].value)]= nodes[i+2][13].value
		fifties_dict[int(nodes[i+2][0].value)]= nodes[i+2][14].value
		sixties_dict[int(nodes[i+2][0].value)]= nodes[i+2][15].value
		seventies_dict[int(nodes[i+2][0].value)]= nodes[i+2][16].value
		eighties_dict[int(nodes[i+2][0].value)]= nodes[i+2][17].value
		nodes_id.append(int(nodes[i+2][0].value))
	print("Adding nodes to the graph...\n")
	G=nx.MultiDiGraph(name="Five Families of New York's graph")
	G.add_nodes_from(nodes_id)
	print("Accessing the Edge-table...")
	workbook_edges = load_workbook(filename="spreadsheets/Edges_list.xlsx")
	workbook_edges.sheetnames
	edges=workbook_edges.active
	print("Adding edges to the graph...\n")
	for i in range(len(edges["A"])-1):
		G.add_edge(edges[i+2][0].value,edges[i+2][1].value, weight=float(edges[i+2][3].value))
	nx.set_node_attributes(G,Name_dict,"Name")
	nx.set_node_attributes(G,Nickname_dict,"Nickname")
	nx.set_node_attributes(G,Boss_dict,"Boss")
	nx.set_node_attributes(G,Acting_Boss_dict,"Acting Boss")
	nx.set_node_attributes(G,Underboss_dict,"Underboss")
	nx.set_node_attributes(G,Acting_Underboss_dict,"Acting Underboss")
	nx.set_node_attributes(G,Consigliere_dict,"Consigliere")
	nx.set_node_attributes(G,Acting_Consigliere_dict,"Acting Consigliere")
	nx.set_node_attributes(G,Capo_dict,"Capo")
	nx.set_node_attributes(G,Family_dict,"Family")
	nx.set_node_attributes(G,twenties_dict,"20's")
	nx.set_node_attributes(G,thirties_dict,"30's")
	nx.set_node_attributes(G,forties_dict,"40's")
	nx.set_node_attributes(G,fifties_dict,"50's")
	nx.set_node_attributes(G,sixties_dict,"60's")
	nx.set_node_attributes(G,seventies_dict,"70's")
	nx.set_node_attributes(G,eighties_dict,"80's")
	print("Graph completed.\n")
	return G,m



def analysis(g):
	print("\nSelect one the following or, if you want to exit, digit 0.\n")
	print("1  - Show the graph's basic information.","\n2  - Show the graph's degree distribution.","\n3  - Show every node's name and family.","\n4  - Show the graph's maximum clique.",
	"\n5  - Draw the graph (Spring-layout).","\n6  - Show the graph's chromatic number.","\n7  - Show the graph's most central characters by degree centrality.",
	"\n8  - Show the graph's most central characters by closeness centrality.","\n9  - Show the graph's degree correlation.",
	"\n10 - Show the graph's critical threshold.\n")
	q = int(input("\n"))
	if q == 1:
		print(basic_information(g))
		time.sleep(8)
		return analysis(g)
	elif q == 2:
		degree_distribution(g)
		return analysis(g)
	elif q == 3:
		print("")
		for n in g.nodes():
			print(n,(4-len(str(n)))*" ", g.nodes[n]["Name"],(25-len(g.nodes[n]["Name"]))*" ", g.nodes[n]["Family"])
		time.sleep(3)
		return analysis(g)
	elif q == 4:
		res=maximum_clique(g)
		print("The maximum clique: ",res[0],"\n")
		time.sleep(3)
		r=int(input("If you want to visualize this clique, digit 1\n"))
		if r==1:
			labels={}
			color=[]
			for i in res[0]:
				labels[i]=g.nodes[i]["Name"]
				color.append(color_map[i])
			ax = plt.gca()
			ax.set_title('Maximum Clique')
			nx.draw(res[1],labels=labels, node_color=color,with_labels = True,ax=ax)
			plt.show()
		return analysis(g)
	elif q == 5:
		labels={}
		for i in g.nodes():
			labels[i]=g.nodes[i]["Name"]
		pos = nx.spring_layout(g,k=0.75)
		with_labels=False
		q = int(input("\nDigit 1 if you want to see each node's name. (Useful if a zoom-in is needed, otherwise it is confusing).\n"))
		if q==1:
			with_labels=True
		nx.draw(g,pos,node_color=color_map,node_size=40,labels=labels,with_labels=with_labels)
		plt.show()
		return analysis(g)
	elif q == 6:
		print(f"\nThe minimum number of colors needed so that adjacent vertices are colored differently is {greedy_colors(g)}.")
		time.sleep(5)
		return analysis(g)
	elif q == 7:
		return central_characters(g)
	elif q == 8:
		return closeness_centrality(g)
	elif q == 9:
		return degree_correlation(g)
	elif q == 10:
		return random_failures(g)
	elif q == 0:
		printgood("Network Analytics: The Five Families in a Net")
		print("\nGianluca Mecci\nMihaela Nedkova Naydenova\nTiziano Iannaccio\nSG")
		return
	else:
		print("\nInvalid choice\n")
		return(analysis(g))



def basic_information(g):
	print("")
	return nx.info(g)



def degree_distribution(g):
	k_i=nx.degree(g)
	d=[]
	for i in k_i:
		d.append(i[1])
	plt.hist(d,color="purple")
	plt.xlabel("Degree",fontsize=12,fontweight="bold")
	plt.ylabel("Number of Nodes",fontsize=12,fontweight="bold")
	plt.title("Degree Distribution",color="purple",fontweight="bold",fontsize=20)
	plt.show()
	q = int(input("\n aDigit 1 if you also want to visualize the probability distribution.\n"))
	if q == 1:
		return probability_distribution(d)
	else:
		return



def probability_distribution(g):
	p=[]
	for i in g:
		p.append(g.count(i)/len(g))
	plt.scatter(g,p,color="red")
	plt.loglog(base=10)
	plt.xlabel("Degree K",fontsize=12,fontweight="bold")
	plt.ylabel("Probability of node i of having degree K",fontsize=12,fontweight="bold")
	plt.title("Probability Degree Distribution",color="red",fontweight="bold",fontsize=18)
	plt.show()
	return



def maximum_clique(g):
	print("This operation will take a few seconds, please wait...\n")
	omega=clique.max_clique(g)
	s=g.subgraph(omega)	
	return omega,s



def greedy_colors(g):
	used_colors=[]
	chi=nx.greedy_color(g)
	for i in chi:
		used_colors.append(chi[i])
	return max(used_colors)



def central_characters(g):
	central_characters = sorted(nx.degree_centrality(g).items(),key= lambda x:x[1], reverse = True)[0:10]
	ids=[]
	for i in central_characters:
		ids.append(i[0])
	y = [central_characters[i][1] for i in range(0,10) ]
	positions=[1,2,3,4,5,6,7,8,9,10]
	labels = []
	barcolor=[]
	for i in ids:
		labels.append(g.nodes[i]["Name"])
		barcolor.append(color_map[i])
	print("These are the ten most central characters of the network (rnked by degree centrality):\n",f"1  - {labels[0]}\n",f"2  - {labels[1]}\n",
	f"3  - {labels[2]}\n",f"4  - {labels[3]}\n",f"5  - {labels[4]}\n",f"6  - {labels[5]}\n",f"7  - {labels[6]}\n",
	f"8  - {labels[7]}\n",f"9  - {labels[8]}\n",f"10 - {labels[9]}\n")
	time.sleep(2)
	plt.figure(figsize=(18, 5))  
	plt.xticks(positions, labels,weight="bold")
	plt.title("Most central characters",color="red",size=25,weight="bold")
	plt.bar(positions,y,width=0.5,align="center",color=barcolor)
	plt.show()
	return analysis(g)



def closeness_centrality(g):
	central_characters = sorted(nx.closeness_centrality(g).items(),key= lambda x:x[1], reverse = True)[0:10]
	ids=[]
	for i in central_characters:
		ids.append(i[0])
	y = [central_characters[i][1] for i in range(0,10) ]
	positions=[1,2,3,4,5,6,7,8,9,10]
	labels = []
	barcolor=[]
	for i in ids:
		labels.append(g.nodes[i]["Name"])
		barcolor.append(color_map[i])
	print("These are the ten most central characters of the network (ranked by closeness centrality):\n",f"1  - {labels[0]}\n",f"2  - {labels[1]}\n",
	f"3  - {labels[2]}\n",f"4  - {labels[3]}\n",f"5  - {labels[4]}\n",f"6  - {labels[5]}\n",f"7  - {labels[6]}\n",
	f"8  - {labels[7]}\n",f"9  - {labels[8]}\n",f"10 - {labels[9]}\n")
	time.sleep(2)
	plt.figure(figsize=(18, 5))  
	plt.xticks(positions, labels,weight="bold")
	plt.title("Most central characters",color="red",size=25,weight="bold")
	plt.bar(positions,y,width=0.5,align="center",color=barcolor)
	plt.show()
	return analysis(g)



def degree_correlation(g):
	kn=nx.k_nearest_neighbors(g)
	knn=list(kn.values())
	k=list(kn.keys())
	plt.title("Degree Correlation",color="red",size=25,weight="bold")
	plt.xlabel("k",fontsize=12,fontweight="bold")
	plt.ylabel(r"$K_{nn}(k)$",fontsize=12,fontweight="bold")
	plt.loglog(k,knn,"go")
	plt.show()
	for i in range(len(knn)):
		knn[i]=math.log10(knn[i])
		k[i]=math.log10(k[i])
	model = lr.LinearRegression().fit(np.array(k).reshape(-1,1),knn)
	print("\nGoodness of fit:",round(model.score(np.array(k).reshape(-1,1),knn),3))
	mu=round(model.coef_[0],3)
	print("Coefficient \u03BC:",mu)
	a=None
	if mu>0.25:
		a="Assortative"
	elif 0.15<mu<=0.35:
		a="Neutral-Assortative"
	elif -0.35<=mu<-0.15:
		a="Neutral-Disassortative"
	elif mu<-0.35:
		a="Disassortative"
	else:
		a="Neutral" 
	print("Type of network:",a)
	time.sleep(2)
	return analysis(g)



def random_failures(g):
	k=[n for d,n in nx.degree(g)]
	k2=[n**2 for d,n in nx.degree(g)]
	m2=np.mean(k2)
	kmr=m2/np.mean(k)
	fc=1-1/(kmr-1)
	print("\nCritical threshold:",round(fc,3))
	time.sleep(2)
	return analysis(g)
	
	
	
start() #starts the program


